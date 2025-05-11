from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook, Workbook
import time
from datetime import datetime
import os

# acessa o site
navegador = webdriver.Chrome()
navegador.get("https://www.google.com/search?q=previsão+do+tempo+sp")
time.sleep(20)

# parametros de clima
temperatura = navegador.find_element(By.ID, 'wob_tm').text
clima = navegador.find_element(By.ID, 'wob_dc').text
umidade = navegador.find_element(By.ID, 'wob_hm').text
vento = navegador.find_element(By.ID, 'wob_ws').text

# exibe as informações do clima
print(f'Temperatura: {temperatura} °C')
print(f'Hoje o clima está {clima}')
print(f'Umidade do ar: {umidade}')
print(f'Velocidade do vento: {vento}')

# abre o arquivo
arquivo_excel = 'clima.xlsx'

# verifica se o arquivo existe
if os.path.exists(arquivo_excel):
    arquivo = load_workbook(arquivo_excel)
    if 'plan' in arquivo.sheetnames:
        plan = arquivo['plan']
    else:
        plan = arquivo.create_sheet('plan')
else:
    arquivo = Workbook()
    plan = arquivo.active
    plan.title = 'plan'
    plan.append(['Data/Hora', 'Temperatura (°C)', 'Clima', 'Umidade', 'Vento'])

# insere os dados na planilha
data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
plan.append([data_hora, temperatura, clima, umidade, vento])
arquivo.save(arquivo_excel)
navegador.quit()
